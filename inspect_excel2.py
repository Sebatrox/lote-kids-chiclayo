import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\Sebastian\Desktop\tienda-ropa\Inventario_Ropa_niños_COMPLETO.xlsx')
ws = wb['Hoja 1']

# Mostrar las primeras 5 filas completas (todas las columnas)
print("Filas 1-6, todas las columnas:")
for row in ws.iter_rows(min_row=1, max_row=6, values_only=True):
    print([repr(v) for v in row])

# Buscar en toda la fila 1 si hay texto que mencione precio/venta/liquidacion
print("\nBusqueda en toda la hoja por 'liquidaci' o 'venta' (primeras 10 filas):")
for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
    for v in row:
        if v and isinstance(v, str) and any(x in v.lower() for x in ['liquidac','venta','precio']):
            print(f"  Encontrado: {repr(v)}")
