import openpyxl, pandas as pd

wb = openpyxl.load_workbook(r'c:\Users\Sebastian\Desktop\tienda-ropa\Inventario_Ropa_niños_COMPLETO.xlsx')
print('Hojas:', wb.sheetnames)

for sh in wb.sheetnames:
    ws = wb[sh]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f'\nHoja "{sh}" ({ws.max_row} filas, {ws.max_column} cols):')
    for i, h in enumerate(headers):
        print(f'  [{i}] {repr(h)}')

# Leer con pandas todas las hojas
print('\n--- Primeras filas de cada hoja ---')
for sh in wb.sheetnames:
    df = pd.read_excel(
        r'c:\Users\Sebastian\Desktop\tienda-ropa\Inventario_Ropa_niños_COMPLETO.xlsx',
        sheet_name=sh
    )
    print(f'\nHoja: {sh}')
    print(df.head(3).to_string())

# Verificar cantidades > 1
df_main = pd.read_excel(r'c:\Users\Sebastian\Desktop\tienda-ropa\Inventario_Ropa_niños_COMPLETO.xlsx')
mask = df_main['Cantidad'] > 1
print(f'\nProductos con Cantidad > 1: {mask.sum()}')
print(df_main[mask][['Codigo','Tipo','Cantidad']].to_string())
